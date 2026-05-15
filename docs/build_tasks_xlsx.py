#!/usr/bin/env python3
"""Generate docs/tasks-done.xlsx — completed tasks with priority + core flag.

Bold rows: priority=High OR core=Yes.
Run: python3 docs/build_tasks_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/tasks-done.xlsx"

# (#, Task, Area, Priority, Core, Status, Commit, Notes)
TASKS = [
    # Foundations
    (1,  "Initial Next.js 16 + TS + Tailwind v4 scaffold",            "Foundation",  "High",   "Yes", "Done", "d5b6192", "App Router, strict TS"),
    (2,  "Supabase clients (browser, server, middleware) + types",     "Foundation",  "High",   "Yes", "Done", "71a33a0", "SSR cookies"),
    (3,  "Initial DB schema + RLS (users, categories, listings, etc.)", "Database",   "High",   "Yes", "Done", "938de5f", "Migration 001"),

    # Core UX
    (4,  "Main layout: header, footer, hero, category grid",           "UI",          "High",   "Yes", "Done", "7f7d660", "Brand tokens"),
    (5,  "Listing card + listing detail + image gallery + contact panel", "Listings",  "High",   "Yes", "Done", "637e9da", "Grid + list variants"),
    (6,  "Post-ad multi-step form + image upload + Claude description", "Listings",  "High",   "Yes", "Done", "18a20e8", "Multi-photo"),
    (7,  "Multi-photo upload upgrade",                                 "Listings",    "Medium", "No",  "Done", "c1e01f8", ""),
    (8,  "View count column + UI",                                     "Listings",    "Low",    "No",  "Done", "89f6ff1", "Migration 003"),

    # Messaging
    (9,  "Inbox + chat thread (conversation model)",                   "Messaging",   "High",   "Yes", "Done", "166a100", "RLS-bound"),
    (10, "Realtime messaging + unread badge in header",                "Messaging",   "High",   "Yes", "Done", "e51ab42", "Migration 002"),
    (11, "Mark-as-read flow + UI polish",                              "Messaging",   "Medium", "No",  "Done", "1bb689b", ""),

    # Watchlist
    (12, "Watchlist DB + API + heart button on card",                  "Watchlist",   "High",   "Yes", "Done", "1569650", "POST/DELETE /api/watchlist"),
    (13, "Watchlist polish: filled heart on watchlist page + remove-on-click refresh + denser grid", "Watchlist", "Medium", "No", "Done", "(this session)", "router.refresh() + initialSaved"),

    # AI
    (14, "Claude description generator endpoint",                      "AI",          "High",   "Yes", "Done", "18a20e8", "Server-only"),
    (15, "Unified AI error handling + watchlist hardening",            "AI",          "Medium", "No",  "Done", "e60a033", ""),
    (16, "Search + reporting error handling",                          "AI",          "Medium", "No",  "Done", "f6bf11a", ""),

    # Admin
    (17, "Admin dashboard, pending listings, reports",                 "Admin",       "High",   "Yes", "Done", "9b02bd7", "RLS-gated"),
    (18, "Per-user is_admin flag (replaces ADMIN_EMAIL env var)",      "Admin",       "High",   "Yes", "Done", "89ab706", "Migration 007, UPDATE revoked"),
    (19, "Admin sign-out flow + middleware redirect logic + admin panel button removal", "Admin", "High", "Yes", "Done", "7e08868", "Logged-in admin auto-routed to /admin"),
    (20, "Auth/admin layout brand alignment (teal header + tree SVG everywhere)", "UI", "Medium", "No", "Done", "7579c42", "Consistent across login, admin, main"),

    # Listing lifecycle
    (21, "Listing expiry cron + email notifications + image reorder",  "Lifecycle",   "High",   "Yes", "Done", "cb74169", "Migrations 005, 006"),

    # Search / browse
    (22, "Postcode + radius distance search on browse",                "Search",      "High",   "Yes", "Done", "cc3e489", "P0 from audit"),
    (23, "Subcategory hierarchy + filter + required post-ad picker",   "Search",      "High",   "Yes", "Done", "8f36efe", "Migration 008, P1 from audit"),
    (24, "Post-ad link styling polish",                                "UI",          "Low",    "No",  "Done", "6375181", ""),

    # Misc
    (25, "Favicon synced with gumtree.com",                            "UI",          "Low",    "No",  "Done", "(this session)", "src/app/favicon.ico"),
    (26, "Project overview doc + tasks Excel",                         "Docs",        "Low",    "No",  "Done", "(this session)", "docs/PROJECT_OVERVIEW.md + this file"),
]

HEADERS = ["#", "Task", "Area", "Priority", "Core", "Status", "Commit", "Notes"]

# Styles
TEAL = "0D475C"
CREAM = "F0ECE6"
LIGHT_GREY = "F4F4F5"
PRIORITY_HIGH_FILL = "FDECEE"   # light red
PRIORITY_MED_FILL = "FFF7E6"    # light amber
PRIORITY_LOW_FILL = "F1F5F9"    # light grey-blue
CORE_FILL = "FEF3C7"            # amber for core badge
border_thin = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

wb = Workbook()

# --- Sheet 1: Tasks ---
ws = wb.active
ws.title = "Tasks Done"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
title = ws.cell(row=1, column=1, value="Gumtree UK Clone — Tasks Done")
title.font = Font(bold=True, size=16, color=CREAM)
title.fill = PatternFill("solid", fgColor=TEAL)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Header row
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = Font(bold=True, color=CREAM)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin
ws.row_dimensions[2].height = 22

# Data rows
for i, row in enumerate(TASKS, start=3):
    num, task, area, priority, core, status, commit, notes = row
    is_high = priority == "High"
    is_core = core == "Yes"
    bold = is_high or is_core
    if priority == "High":
        fill_color = PRIORITY_HIGH_FILL
    elif priority == "Medium":
        fill_color = PRIORITY_MED_FILL
    else:
        fill_color = PRIORITY_LOW_FILL

    values = [num, task, area, priority, core, status, commit, notes]
    for col, v in enumerate(values, 1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = Font(bold=bold)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border_thin

    # Core column gets stronger highlight
    if is_core:
        core_cell = ws.cell(row=i, column=5)
        core_cell.fill = PatternFill("solid", fgColor=CORE_FILL)
        core_cell.font = Font(bold=True, color="92400E")
        core_cell.alignment = Alignment(horizontal="center", vertical="center")

# Column widths
widths = [5, 60, 13, 10, 7, 9, 16, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws.freeze_panes = "A3"

# Filter
ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{2+len(TASKS)}"

# --- Sheet 2: Legend ---
ws2 = wb.create_sheet("Legend")
ws2.cell(row=1, column=1, value="Legend").font = Font(bold=True, size=14, color=CREAM)
ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor=TEAL)
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

legend = [
    ("Priority: High",   "Core marketplace function — buying, selling, messaging, search, auth, admin"),
    ("Priority: Medium", "Quality of life or polish on top of a shipped core feature"),
    ("Priority: Low",    "Cosmetic, docs, or small UX wins"),
    ("Core = Yes",       "Feature is part of the minimum-viable marketplace; removing it breaks the product"),
    ("Bold row",         "Either High priority or Core feature — read these first"),
    ("Status: Done",     "Merged to main or in current branch"),
]
for i, (k, v) in enumerate(legend, start=2):
    kc = ws2.cell(row=i, column=1, value=k)
    vc = ws2.cell(row=i, column=2, value=v)
    kc.font = Font(bold=True)
    kc.alignment = Alignment(vertical="center")
    vc.alignment = Alignment(vertical="center", wrap_text=True)
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 80

# --- Sheet 3: Summary ---
ws3 = wb.create_sheet("Summary")
ws3.cell(row=1, column=1, value="Summary").font = Font(bold=True, size=14, color=CREAM)
ws3.cell(row=1, column=1).fill = PatternFill("solid", fgColor=TEAL)
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

total = len(TASKS)
high = sum(1 for t in TASKS if t[3] == "High")
med  = sum(1 for t in TASKS if t[3] == "Medium")
low  = sum(1 for t in TASKS if t[3] == "Low")
core = sum(1 for t in TASKS if t[4] == "Yes")

rows = [
    ("Total tasks shipped", total),
    ("High priority",       high),
    ("Medium priority",     med),
    ("Low priority",        low),
    ("Core features",       core),
]
for i, (k, v) in enumerate(rows, start=2):
    kc = ws3.cell(row=i, column=1, value=k)
    vc = ws3.cell(row=i, column=2, value=v)
    kc.font = Font(bold=True)
    vc.font = Font(bold=True, color=TEAL, size=12)
    vc.alignment = Alignment(horizontal="center")
ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 12

wb.save(OUT)
print(f"Wrote {OUT} — {total} tasks ({high} High, {core} Core)")
